import openpyxl

# defind sheet title
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.title = 'Spam Bacon Eggs Sheet'
# print(wb.get_sheet_names())


# rename sheet title and save
# wb = openpyxl.open('example.xlsx')
# sheet = wb.active
# sheet.title = 'Spam'
# wb.save('example_copy.xlsx')

# work table create and remove
# wb = openpyxl.Workbook()
# wb.create_sheet()
# print(wb.create_sheet(index=0,title='First Sheet'))
# print(wb.create_sheet(index=2,title='Second Sheet'))
# print(wb.get_sheet_names())
# wb.remove_sheet(wb.get_sheet_by_name('Second Sheet'))
# wb.remove_sheet(wb.get_sheet_by_name('Sheet1'))
# print(wb.get_sheet_names())


# value insert the table
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1']='Hello World'
# print(sheet['A1'].value)
# wb.save('res.xlsx')

# ===================更新Excel数据==============================
'''
# update table
wb = openpyxl.open('produceSales.xlsx')
sheet = wb.active

# update the prices
PRICE_UPDATES = {
    'Garlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}

# TODO: Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save('updateProduceSales.xlsx')
'''

# ===================设置单元格的字体风格==============================

from openpyxl.styles import Font

# wb = openpyxl.Workbook()
# sheet = wb.active
# stylefont=Font(size=20,italic=True) #italic斜体
# sheet['A1'].font = stylefont
# sheet['A1']='Hello world'

# font1 = Font(name='Times New Roman', bold=True)  # bold粗体
# sheet['A1'].font = font1
# sheet['A1']='Bold Times'
#
# font2 = Font(size=24, italic=True)
# sheet['B1'].font = font2
# sheet['B1'] = '24 pt Italic'
# wb.save('style.xlsx')


# ===================公式==============================
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1']=200
# sheet['A2']=300
# sheet['A3']='=SUM(A1:A2)'
# wb.save('writeSUM.xlsx')

# wb = openpyxl.open('writeSUM.xlsx')
# sheet = wb.active
# print(sheet['A3'].value)

# 没有保存数据无法使用data_ongly=True来查询公式的结果数据

# ===================调整行和列==============================
#  wb = openpyxl.Workbook()
# sheet = wb.active
# sheet['A1'] = 'Tall row'
# sheet['B2'] = 'Wide column'
# sheet.row_dimensions[1].height = 70  # 设置行高
# sheet.column_dimensions['B'].width = 20  # 设置行宽
# wb.save('dimensions.xlsx')

# 合并、拆分单元格
# wb = openpyxl.Workbook()
# sheet = wb.active
# sheet.merge_cells('A1:D3')
# sheet['A1']='Twelve cells merged together.'
# sheet.merge_cells('C5:D5')
# sheet['C5']='Two merged cells.'
# wb.save('merged.xlsx')

# wb = openpyxl.open('merged.xlsx')
# sheet = wb.active
# sheet.unmerge_cells('A1:D3')
# sheet.unmerge_cells('C5:D5')
# wb.save('merged.xlsx')

# 冻结窗格
# freeze_panes的设置                                       冻结的行和列
# sheet.freeze_panes = 'A2'                                 行1
# sheet.freeze_panes = 'B1'                                 列A
# sheet.freeze_panes = 'C1'                                 列A和列B
# sheet.freeze_panes = 'C2'                                 行1和列A和列B
# sheet.freeze_panes = 'A1'或 sheet.freeze_panes = None      没有冻结窗格

# wb = openpyxl.open('produceSales.xlsx')
# sheet = wb.active
# sheet.freeze_panes = 'C2'
# wb.save('freezeExample.xlsx')


# ===================图表==============================
# wb = openpyxl.Workbook()
# sheet = wb.active
# for i in range(1, 11):
#     # sheet['A' + str(i)] = i
#     sheet.append([i])
# wb.save('SamepleChart.xlsx')

# from openpyxl.chart import BarChart,Reference,Series
# values = Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
# chart = BarChart()
# chart.add_data(values)
# sheet.add_chart(chart, "E15")
# wb.save("SampleChart.xlsx")


'''
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.layout import Layout, ManualLayout

wb = Workbook()
ws = wb.active

rows = [
    ['Size', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 25],
    [6, 25, 35],
    [7, 20, 40],
]

for row in rows:
    ws.append(row)

ch1 = ScatterChart()
xvalues = Reference(ws, min_col=1, min_row=2, max_row=7)
for i in range(2, 4):
    values = Reference(ws, min_col=i, min_row=1, max_row=7)
    series = Series(values, xvalues, title_from_data=True)
    ch1.series.append(series)


ch1.title = "Default layout"
ch1.style = 13
ch1.x_axis.title = 'Size'
ch1.y_axis.title = 'Percentage'
ch1.legend.position = 'r'

ws.add_chart(ch1, "B10")

from copy import deepcopy

# Half-size chart, bottom right
ch2 = deepcopy(ch1)
ch2.title = "Manual chart layout"
ch2.legend.position = "tr"
ch2.layout=Layout(
    manualLayout=ManualLayout(
        x=0.25, y=0.25,
        h=0.5, w=0.5,
    )
)
ws.add_chart(ch2, "H10")

# Half-size chart, centred
ch3 = deepcopy(ch1)
ch3.layout = Layout(
    ManualLayout(
    x=0.25, y=0.25,
    h=0.5, w=0.5,
    xMode="edge",
    yMode="edge",
    )
)
ch3.title = "Manual chart layout, edge mode"
ws.add_chart(ch3, "B27")

# Manually position the legend bottom left
ch4 = deepcopy(ch1)
ch4.title = "Manual legend layout"
ch4.legend.layout = Layout(
    manualLayout=ManualLayout(
        yMode='edge',
        xMode='edge',
        x=0, y=0.9,
        h=0.1, w=0.5
    )
)

ws.add_chart(ch4, "H27")

wb.save("chart_layout.xlsx")
'''

# 乘法表
#
# import sys
# import openpyxl
# from openpyxl.styles import Font
#
# try:
#     num = int(sys.argv[1])
#     wb = openpyxl.Workbook()
#     sheet = wb.active
#     styleFont = Font(bold=True)
#     for i in range(1, num + 1):
#         #首行、首列加粗
#         sheet.cell(row=1, column=i + 1).font = styleFont
#         sheet.cell(row=i + 1, column=1).font = styleFont
#         #首行、首列数据
#         sheet.cell(row=1, column=i + 1).value = i
#         sheet.cell(row=i + 1, column=1).value = i
#
#         for j in range(1, num + 1):
#             sheet.cell(row=i + 1, column=j + 1).value = i * j
#
#     wb.save('res.xlsx')
# except IndexError as err:
#     print('''
#         Please input a the Number:
#         example
#         python   xxxx.py  6
#     ''')


# excel 插入空行
'''
import sys
import openpyxl

try:
    start_row = int(sys.argv[1])
    insert_row = int(sys.argv[2])
    file = sys.argv[3]
    wb = openpyxl.open('res.xlsx')
    sheet = wb.active
    sheet.insert_rows(start_row,insert_row)
    wb.save('res.xlsx')
except Exception as err:
    print('缺少参数如:python  xxx.py 3 2 xxx.xlsx',err)

'''




